"""Multi-engine ingestion service with deterministic integrity reporting."""
from __future__ import annotations

import hashlib
import inspect
import re
import shutil
import uuid
from datetime import datetime, timezone
from collections.abc import Awaitable, Callable
from pathlib import Path
from typing import Any, Optional

import duckdb
import pandas as pd
import polars as pl
from openpyxl import load_workbook

from app.core.config import get_settings
from app.core.exceptions import FileTooLargeException, InconsistencyException, UnsupportedFormatException
from app.core.logging import logger
from app.schemas import ColumnSchema, ColumnType, IntegrityReport, SheetMetadata


settings = get_settings()

SUPPORTED_FORMATS = {".xlsx", ".xls", ".xlsm", ".xlsb", ".csv", ".tsv", ".ods"}
FORMAT_MAP = {
    ".xlsx": "xlsx",
    ".xls": "xls",
    ".xlsm": "xlsm",
    ".xlsb": "xlsb",
    ".csv": "csv",
    ".tsv": "tsv",
    ".ods": "ods",
}

_PERCENT_RE = re.compile(r"^[-+]?\d+(?:[\.,]\d+)?\s*%$")
_CURRENCY_RE = re.compile(r"^(?:R\$|\$|€|£|¥)\s*[-+]?\d+(?:[\.,]\d+)?$")
_INT_RE = re.compile(r"^[-+]?\d+$")
_FLOAT_RE = re.compile(r"^[-+]?\d+[\.,]\d+$")


class IngestionService:
    """Service responsible for reading, validating and indexing tabular files."""

    def __init__(self) -> None:
        self.storage_path = Path(settings.STORAGE_PATH)
        self.storage_path.mkdir(parents=True, exist_ok=True)

    async def process_upload(
        self,
        file_path: str,
        original_filename: str,
        strict_mode: bool = False,
        preferred_date_format: Optional[str] = None,
        encoding: Optional[str] = None,
        delimiter: Optional[str] = None,
        progress_callback: Callable[[str, int, str], Awaitable[None] | None] | None = None,
    ) -> dict[str, Any]:
        """Process a file and produce deterministic metadata and integrity report."""
        del preferred_date_format  # Reserved for future parsing overrides.

        async def emit_progress(stage: str, progress: int, message: str) -> None:
            if progress_callback is None:
                return

            result = progress_callback(stage, progress, message)
            if inspect.isawaitable(result):
                await result

        started_at = datetime.now(tz=timezone.utc)
        source_path = Path(file_path)
        extension = source_path.suffix.lower()

        if extension not in SUPPORTED_FORMATS:
            raise UnsupportedFormatException(f"Formato nao suportado: {extension}")

        file_size = source_path.stat().st_size
        max_size = settings.MAX_FILE_SIZE_MB * 1024 * 1024
        if file_size > max_size:
            raise FileTooLargeException(
                f"Arquivo de {file_size} bytes excede limite de {max_size} bytes"
            )

        await emit_progress("hashing", 8, "Calculando assinatura do arquivo")
        file_hash = await self._calculate_sha256(str(source_path))
        file_uuid = str(uuid.uuid5(uuid.NAMESPACE_URL, file_hash))
        file_format = FORMAT_MAP[extension]

        storage_dir = self.storage_path / file_uuid
        storage_dir.mkdir(parents=True, exist_ok=True)
        stored_original_path = storage_dir / f"original{extension}"
        if not stored_original_path.exists():
            shutil.copy2(source_path, stored_original_path)

        await emit_progress("reading", 25, "Lendo estrutura do arquivo")
        primary = await self._read_primary(stored_original_path, file_format, encoding, delimiter)
        engines_used = ["polars/calamine"]
        secondary: dict[str, dict[str, Any]] = {}

        if file_format in {"xlsx", "xlsm"}:
            await emit_progress("reading_secondary", 40, "Comparando leitura com outra engine")
            secondary = await self._read_secondary_xlsx(stored_original_path)
            engines_used.append("openpyxl")

        if file_format == "xlsb":
            engines_used.append("pyxlsb")

        await emit_progress("validating", 58, "Validando consistencia e qualidade dos dados")
        divergences = self._compare_readings(primary, secondary)

        await emit_progress("profiling", 72, "Gerando esquema e perfil das abas")
        sheets_metadata: list[SheetMetadata] = []
        for index, (sheet_name, frame) in enumerate(primary.items()):
            sheet_hash = self._hash_polars_sheet(frame)
            sheets_metadata.append(self._analyze_sheet(frame, sheet_name, index, sheet_hash))

        secondary_totals = self._secondary_totals(secondary)
        report = self._generate_integrity_report(
            file_hash=file_hash,
            sheets=sheets_metadata,
            engines_used=engines_used,
            integrity_issues=divergences,
            strict_mode=strict_mode,
            formulas_detected=secondary_totals["formulas_detected"],
            merged_cells_detected=secondary_totals["merged_cells_detected"],
        )

        if strict_mode and divergences:
            report.warnings.append(
                "Modo estrito detectou divergencias, mas o dataset foi persistido como inconsistente."
            )

        await emit_progress("persisting", 88, "Persistindo dados analiticos")
        duckdb_path = await self._persist_to_duckdb(file_uuid, primary)
        status = "inconsistent" if divergences else "completed"
        elapsed_seconds = (datetime.now(tz=timezone.utc) - started_at).total_seconds()

        await emit_progress("finalizing", 100, "Processamento concluido")

        return {
            "uuid": file_uuid,
            "original_filename": original_filename,
            "file_format": file_format,
            "file_size_bytes": file_size,
            "file_hash_sha256": file_hash,
            "status": status,
            "strict_mode": strict_mode,
            "total_sheets": len(sheets_metadata),
            "total_rows": sum(sheet.row_count for sheet in sheets_metadata),
            "total_columns": max((sheet.column_count for sheet in sheets_metadata), default=0),
            "sheets": [sheet.model_dump() for sheet in sheets_metadata],
            "integrity_report": report.model_dump(),
            "duckdb_path": str(duckdb_path),
            "storage_path": str(stored_original_path),
            "processing_time_seconds": elapsed_seconds,
        }

    async def _calculate_sha256(self, file_path: str) -> str:
        hash_obj = hashlib.sha256()
        with open(file_path, "rb") as file_obj:
            for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
                hash_obj.update(chunk)
        return hash_obj.hexdigest()

    async def _read_primary(
        self,
        file_path: Path,
        file_format: str,
        encoding: Optional[str],
        delimiter: Optional[str],
    ) -> dict[str, pl.DataFrame]:
        if file_format in {"csv", "tsv"}:
            sep = delimiter or ("\t" if file_format == "tsv" else ",")
            detected_encoding = encoding or self._detect_encoding(file_path)
            frame = pl.read_csv(
                file_path,
                separator=sep,
                encoding=detected_encoding,
                infer_schema_length=0,
                null_values=["", "NULL", "null", "NA", "N/A", "#N/A"],
            )
            return {"data": self._normalize_tabular_frame(frame)}

        if file_format in {"xlsx", "xls", "xlsm", "ods"}:
            if file_format in {"xlsx", "xlsm"}:
                workbook = load_workbook(file_path, read_only=True, data_only=False)
                sheet_names = list(workbook.sheetnames)
                workbook.close()
            elif file_format == "xls":
                # .xls (BIFF) nao e compatível com openpyxl; usa xlrd para listar abas
                try:
                    excel = pd.ExcelFile(file_path, engine="xlrd")
                    sheet_names = list(excel.sheet_names)
                except Exception:
                    sheet_names = ["data"]
            else:
                sheet_names = ["data"]

            sheets: dict[str, pl.DataFrame] = {}
            for sheet_name in sheet_names:
                try:
                    sheets[sheet_name] = pl.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        engine="calamine",
                        infer_schema_length=0,
                    )
                except Exception:
                    engine = "xlrd" if file_format == "xls" else "openpyxl"
                    fallback = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine, dtype=str)
                    sheets[sheet_name] = pl.from_pandas(fallback.fillna(value=""))
            return {name: self._normalize_tabular_frame(df) for name, df in sheets.items()}

        if file_format == "xlsb":
            frame_pd = pd.read_excel(file_path, engine="pyxlsb", dtype=str)
            return {"data": self._normalize_tabular_frame(pl.from_pandas(frame_pd.fillna(value="")))}

        raise UnsupportedFormatException(f"Formato nao implementado para leitura primaria: {file_format}")

    def _normalize_tabular_frame(self, frame: pl.DataFrame) -> pl.DataFrame:
        """Promove linha de cabecalho quando o reader gera colunas Unnamed:* (Excel com titulos/linhas em branco)."""
        if frame.width == 0 or frame.height == 0:
            return frame

        columns = [str(c) for c in frame.columns]
        unnamed = sum(1 for name in columns if name.lower().startswith("unnamed"))
        if unnamed / max(len(columns), 1) < 0.6:
            return frame

        header_idx = self._detect_header_row_index(frame)
        if header_idx is None:
            return frame

        header_values = frame.row(header_idx)
        new_names: list[str] = []
        for idx, value in enumerate(header_values):
            text = str(value).strip() if value is not None else ""
            if text == "" or text.lower() == "nan":
                new_names.append(f"col_{idx}")
            else:
                new_names.append(text)

        new_names = self._dedupe_column_names(new_names)

        trimmed = frame.slice(header_idx + 1, max(0, frame.height - (header_idx + 1)))
        if trimmed.height == 0:
            return frame

        trimmed = trimmed.rename({old: new for old, new in zip(trimmed.columns, new_names)})
        return trimmed

    def _detect_header_row_index(self, frame: pl.DataFrame) -> Optional[int]:
        max_scan = min(15, frame.height)
        best_idx: Optional[int] = None
        best_score = -1.0

        for idx in range(max_scan):
            row = frame.row(idx)
            texts: list[str] = []
            for value in row:
                if value is None:
                    texts.append("")
                    continue
                text = str(value).strip()
                if text.lower() == "nan":
                    text = ""
                texts.append(text)

            non_empty = [t for t in texts if t != ""]
            if len(non_empty) < 3:
                continue

            unique_ratio = len(set(non_empty)) / len(non_empty)
            fill_ratio = len(non_empty) / max(len(texts), 1)

            numeric_hits = 0
            for t in non_empty[: min(30, len(non_empty))]:
                if _INT_RE.match(t) or _FLOAT_RE.match(t):
                    numeric_hits += 1
            numeric_ratio = numeric_hits / max(min(30, len(non_empty)), 1)

            score = (unique_ratio * 1.2) + (fill_ratio * 0.8) - (numeric_ratio * 1.1)
            if score > best_score:
                best_score = score
                best_idx = idx

        return best_idx

    def _dedupe_column_names(self, names: list[str]) -> list[str]:
        seen: dict[str, int] = {}
        result: list[str] = []
        for name in names:
            base = name or "col"
            count = seen.get(base, 0)
            final_name = base if count == 0 else f"{base}_{count}"
            seen[base] = count + 1
            result.append(final_name)
        return result

    async def _read_secondary_xlsx(self, file_path: Path) -> dict[str, dict[str, Any]]:
        workbook = load_workbook(file_path, read_only=False, data_only=False)
        result: dict[str, dict[str, Any]] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            effective_rows, effective_columns = self._get_effective_sheet_bounds(worksheet)
            result[sheet_name] = {
                "max_row": effective_rows,
                "max_column": effective_columns,
                "merged_cells_count": len(worksheet.merged_cells.ranges),
                "formula_count": self._count_formulas(worksheet),
                "sheet_hash": self._hash_sheet_structure(worksheet),
            }
        workbook.close()
        return result

    def _get_effective_sheet_bounds(self, worksheet: Any) -> tuple[int, int]:
        last_non_empty_row = 0
        last_non_empty_column = 0

        for row_index, row in enumerate(worksheet.iter_rows(values_only=False), start=1):
            row_has_value = False
            for column_index, cell in enumerate(row, start=1):
                value = cell.value
                if value in (None, ""):
                    continue

                row_has_value = True
                if column_index > last_non_empty_column:
                    last_non_empty_column = column_index

            if row_has_value:
                last_non_empty_row = row_index

        return max(0, last_non_empty_row - 1), last_non_empty_column

    def _compare_readings(
        self,
        primary: dict[str, pl.DataFrame],
        secondary: dict[str, dict[str, Any]],
    ) -> list[dict[str, Any]]:
        if not secondary:
            return []

        divergences: list[dict[str, Any]] = []
        primary_sheets = set(primary.keys())
        secondary_sheets = set(secondary.keys())

        if primary_sheets != secondary_sheets:
            divergences.append(
                {
                    "type": "sheet_set_mismatch",
                    "severity": "high",
                    "primary_sheets": sorted(primary_sheets),
                    "secondary_sheets": sorted(secondary_sheets),
                }
            )

        for sheet_name in sorted(primary_sheets & secondary_sheets):
            frame = primary[sheet_name]
            secondary_meta = secondary[sheet_name]

            if frame.height != int(secondary_meta["max_row"]):
                divergences.append(
                    {
                        "type": "row_count_mismatch",
                        "severity": "medium",
                        "sheet": sheet_name,
                        "primary_rows": frame.height,
                        "secondary_rows": int(secondary_meta["max_row"]),
                    }
                )

            if frame.width != int(secondary_meta["max_column"]):
                divergences.append(
                    {
                        "type": "column_count_mismatch",
                        "severity": "medium",
                        "sheet": sheet_name,
                        "primary_columns": frame.width,
                        "secondary_columns": int(secondary_meta["max_column"]),
                    }
                )

        return divergences

    def _analyze_sheet(
        self,
        dataframe: pl.DataFrame,
        sheet_name: str,
        sheet_index: int,
        sheet_hash: str,
    ) -> SheetMetadata:
        columns: list[ColumnSchema] = []
        for index, column_name in enumerate(dataframe.columns):
            series = dataframe[column_name]
            detected_type = self._detect_column_type(series)
            sample_values = [str(value) for value in series.drop_nulls().head(10).to_list()]
            columns.append(
                ColumnSchema(
                    name=str(column_name),
                    index=index,
                    detected_type=detected_type,
                    null_count=int(series.null_count()),
                    unique_count=int(series.n_unique()),
                    cardinality=int(series.n_unique()),
                    sample_values=sample_values,
                )
            )

        return SheetMetadata(
            name=sheet_name,
            index=sheet_index,
            row_count=dataframe.height,
            column_count=dataframe.width,
            columns=columns,
            sheet_hash=sheet_hash,
        )

    def _detect_column_type(self, col: pl.Series) -> ColumnType:
        dtype = col.dtype

        if dtype in (pl.Int8, pl.Int16, pl.Int32, pl.Int64, pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64):
            return "integer"
        if dtype in (pl.Float32, pl.Float64):
            return "float"
        if dtype == pl.Decimal:
            return "decimal"
        if dtype == pl.Boolean:
            return "boolean"
        if dtype == pl.Date:
            return "date"
        if dtype == pl.Datetime:
            return "datetime"

        values = [str(value).strip() for value in col.drop_nulls().to_list() if str(value).strip() != ""]
        return self._infer_string_type(values)

    def _infer_string_type(self, samples: list[str]) -> ColumnType:
        if not samples:
            return "empty"

        lowered = [value.lower() for value in samples]
        total = len(samples)

        if all(_PERCENT_RE.match(value) for value in samples):
            return "percentage"
        if all(_CURRENCY_RE.match(value) for value in samples):
            return "currency"

        bool_values = {"true", "false", "yes", "no", "sim", "nao", "1", "0"}
        if all(value in bool_values for value in lowered):
            return "boolean"

        has_leading_zero_numeric = any(
            len(value) > 1 and value.startswith("0") and value[1:].isdigit() for value in samples
        )
        if has_leading_zero_numeric:
            return "string"

        integer_hits = sum(1 for value in samples if _INT_RE.match(value))
        float_hits = sum(1 for value in samples if _FLOAT_RE.match(value))
        date_hits = sum(1 for value in samples if self._looks_like_date(value))

        matched_families = sum(1 for hits in (integer_hits, float_hits, date_hits) if hits > 0)
        if matched_families >= 2 and (integer_hits + float_hits + date_hits) / total > 0.9:
            return "mixed"

        if integer_hits == total:
            return "integer"
        if integer_hits + float_hits == total and float_hits > 0:
            return "float"
        if date_hits / total >= 0.9:
            return "date"

        unique_count = len(set(lowered))
        if unique_count <= 50 and (unique_count / total) <= 0.2:
            return "categorical"

        return "string"

    async def _persist_to_duckdb(self, uuid_str: str, sheets_data: dict[str, pl.DataFrame]) -> Path:
        db_path = self.storage_path / uuid_str / "analytics.db"
        connection = duckdb.connect(str(db_path))
        try:
            for sheet_name, dataframe in sheets_data.items():
                table_name = self._sanitize_table_name(sheet_name)
                arrow_table = dataframe.to_arrow()
                connection.register("temp_sheet", arrow_table)
                connection.execute(f"DROP TABLE IF EXISTS {table_name}")
                connection.execute(f"CREATE TABLE {table_name} AS SELECT * FROM temp_sheet")
                connection.unregister("temp_sheet")

                for column_name in dataframe.columns[:3]:
                    index_name = f"idx_{table_name}_{self._sanitize_table_name(column_name)}"
                    connection.execute(
                        f'CREATE INDEX IF NOT EXISTS {index_name} ON {table_name}("{column_name}")'
                    )
        finally:
            connection.close()
        return db_path

    def _sanitize_table_name(self, name: str) -> str:
        sanitized = "".join(char if char.isalnum() else "_" for char in name)
        return f"sheet_{sanitized}"[:60]

    def _generate_integrity_report(
        self,
        file_hash: str,
        sheets: list[SheetMetadata],
        engines_used: list[str],
        integrity_issues: list[dict[str, Any]],
        strict_mode: bool,
        formulas_detected: int,
        merged_cells_detected: int,
    ) -> IntegrityReport:
        columns_by_type: dict[str, int] = {}
        mixed_columns: list[str] = []
        for sheet in sheets:
            for column in sheet.columns:
                column_type = column.detected_type
                columns_by_type[column_type] = columns_by_type.get(column_type, 0) + 1
                if column_type == "mixed":
                    mixed_columns.append(f"{sheet.name}.{column.name}")

        warnings: list[str] = []
        errors: list[str] = []
        for issue in integrity_issues:
            message = f"{issue.get('type', 'unknown')} ({issue.get('sheet', 'global')})"
            if issue.get("severity") == "high":
                errors.append(message)
            else:
                warnings.append(message)

        return IntegrityReport(
            total_sheets=len(sheets),
            total_rows=sum(sheet.row_count for sheet in sheets),
            total_columns=max((sheet.column_count for sheet in sheets), default=0),
            cells_read=sum(sheet.row_count * sheet.column_count for sheet in sheets),
            empty_cells=sum(column.null_count for sheet in sheets for column in sheet.columns),
            columns_by_type=columns_by_type,
            mixed_type_columns=mixed_columns,
            formulas_detected=formulas_detected,
            merged_cells_detected=merged_cells_detected,
            file_hash=file_hash,
            file_hash_sha256=file_hash,
            sheet_hashes={sheet.name: sheet.sheet_hash for sheet in sheets},
            engines_used=engines_used,
            engine_divergences=integrity_issues,
            warnings=warnings,
            errors=errors,
            strict_mode_blocked=strict_mode and bool(integrity_issues),
        )

    def _detect_encoding(self, file_path: Path) -> str:
        try:
            import chardet

            with open(file_path, "rb") as file_obj:
                detection = chardet.detect(file_obj.read(100000))
            return detection.get("encoding") or "utf-8"
        except Exception:
            return "utf-8"

    def _secondary_totals(self, secondary: dict[str, dict[str, Any]]) -> dict[str, int]:
        return {
            "formulas_detected": sum(int(item.get("formula_count", 0)) for item in secondary.values()),
            "merged_cells_detected": sum(int(item.get("merged_cells_count", 0)) for item in secondary.values()),
        }

    def _looks_like_date(self, value: str) -> bool:
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
        return not pd.isna(parsed)

    def _count_formulas(self, worksheet: Any) -> int:
        total = 0
        for row in worksheet.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total += 1
        return total

    def _hash_sheet_structure(self, worksheet: Any) -> str:
        signature = f"{worksheet.max_row}|{worksheet.max_column}|{len(worksheet.merged_cells.ranges)}"
        return hashlib.sha256(signature.encode("utf-8")).hexdigest()

    def _hash_polars_sheet(self, dataframe: pl.DataFrame) -> str:
        payload = dataframe.head(1000).to_dicts()
        return hashlib.sha256(str(payload).encode("utf-8")).hexdigest()


ingestion_service = IngestionService()
