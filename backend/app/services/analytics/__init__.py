"""Serviço de analytics com heurísticas determinísticas para sugestão de gráficos.

TODA a lógica é determinística e auditável. Nenhuma IA é utilizada.
"""
from datetime import datetime
from collections import OrderedDict
from pathlib import Path
import re
import unicodedata
from typing import Any, Optional

import duckdb
import pandas as pd
import polars as pl
from openpyxl import load_workbook

from app.core.config import get_settings
from app.core.logging import logger
from app.schemas import ChartSuggestion, ColumnSchema, ColumnType


settings = get_settings()

_PERCENT_RE = re.compile(r"^[-+]?\d+(?:[\.,]\d+)?\s*%$")
_CURRENCY_RE = re.compile(r"^(?:R\$|\$|€|£|¥)\s*[-+]?\d+(?:[\.,]\d+)?$")
_INT_RE = re.compile(r"^[-+]?\d+$")
_FLOAT_RE = re.compile(r"^[-+]?\d+[\.,]\d+$")


class AnalyticsService:
    """Serviço de análise e sugestão de visualizações."""

    def __init__(self):
        self.storage_path = Path(settings.STORAGE_PATH)

    def suggest_charts(self, file_uuid: str, sheet_name: str) -> list[ChartSuggestion]:
        """Sugere gráficos baseado em heurísticas determinísticas."""

        db_path = self.storage_path / file_uuid / "analytics.db"
        if not db_path.exists():
            return []

        conn = duckdb.connect(str(db_path), read_only=True)
        table_name = self._sanitize_table_name(sheet_name)

        try:
            schema_df = conn.execute(f"DESCRIBE {table_name}").fetchdf()
            columns = self._extract_column_info(conn, table_name, schema_df)
        except Exception as e:
            logger.error(f"Erro ao obter schema: {e}")
            conn.close()
            return []

        conn.close()

        categorical_cols = [
            c for c in columns
            if c.detected_type in ("categorical", "string", "boolean")
            and not self._is_generic_column_name(c.name)
            and not self._is_month_column_name(c.name)
            and not self._is_numeric_metric_candidate(c)
        ]
        numeric_cols = [
            c for c in columns
            if c.detected_type in ("integer", "float", "decimal", "currency", "percentage") and not self._is_generic_column_name(c.name)
        ]
        date_cols = [c for c in columns if c.detected_type in ("date", "datetime")]

        suggestions: list[ChartSuggestion] = []
        seen_suggestions: set[tuple[str, tuple[str, ...], tuple[str, ...], str]] = set()

        for suggestion in self._build_financial_suggestions(columns):
            self._append_suggestion(suggestions, seen_suggestions, suggestion)

        # HEURISTICA 1: 1 dimensao categorica + 1 metrica = barras
        if categorical_cols and numeric_cols:
            for cat in categorical_cols[:3]:
                for num in numeric_cols[:3]:
                    if cat.cardinality <= 50:
                        self._append_suggestion(
                            suggestions,
                            seen_suggestions,
                            ChartSuggestion(
                                chart_type="bar",
                                title=f"{num.name} por {cat.name}",
                                description=f"Comparacao de {num.name} agrupado por {cat.name}",
                                dimension_columns=[cat.name],
                                metric_columns=[num.name],
                                confidence_score=0.95 if cat.cardinality <= 20 else 0.75,
                                heuristic_rule="categorical_1_metric_1: 1 dimensao categorica (card <=50) + 1 metrica -> bar chart",
                                recommended_aggregation="sum",
                            ),
                        )

        # HEURISTICA 2: 1 data + 1+ metricas = linha ou area
        if date_cols and numeric_cols:
            for date_col in date_cols[:2]:
                for num in numeric_cols[:2]:
                    self._append_suggestion(
                        suggestions,
                        seen_suggestions,
                        ChartSuggestion(
                            chart_type="line",
                            title=f"Evolucao de {num.name} ao longo do tempo",
                            description=f"Serie temporal de {num.name} por {date_col.name}",
                            dimension_columns=[date_col.name],
                            metric_columns=[num.name],
                            confidence_score=0.92,
                            heuristic_rule="temporal_1_metric_1: 1 coluna temporal + 1 metrica -> line chart",
                            recommended_aggregation="sum",
                        ),
                    )
                    self._append_suggestion(
                        suggestions,
                        seen_suggestions,
                        ChartSuggestion(
                            chart_type="area",
                            title=f"Area de {num.name} ao longo do tempo",
                            description=f"Visualizacao de area de {num.name} por {date_col.name}",
                            dimension_columns=[date_col.name],
                            metric_columns=[num.name],
                            confidence_score=0.85,
                            heuristic_rule="temporal_1_metric_1_area: 1 coluna temporal + 1 metrica -> area chart",
                            recommended_aggregation="sum",
                        ),
                    )

        # HEURISTICA 3: Categoria baixa cardinalidade + metrica = pizza/donut
        if categorical_cols and numeric_cols:
            for cat in categorical_cols:
                if 2 <= cat.cardinality <= 8:
                    for num in numeric_cols[:1]:
                        self._append_suggestion(
                            suggestions,
                            seen_suggestions,
                            ChartSuggestion(
                                chart_type="pie",
                                title=f"Distribuicao de {num.name} por {cat.name}",
                                description=f"Proporcao de {num.name} entre categorias de {cat.name}",
                                dimension_columns=[cat.name],
                                metric_columns=[num.name],
                                confidence_score=0.90,
                                heuristic_rule="low_cardinality_pie: categoria com cardinalidade 2-8 + 1 metrica -> pie chart",
                                recommended_aggregation="sum",
                            ),
                        )
                        self._append_suggestion(
                            suggestions,
                            seen_suggestions,
                            ChartSuggestion(
                                chart_type="donut",
                                title=f"Distribuicao de {num.name} por {cat.name}",
                                description=f"Proporcao de {num.name} entre categorias de {cat.name}",
                                dimension_columns=[cat.name],
                                metric_columns=[num.name],
                                confidence_score=0.88,
                                heuristic_rule="low_cardinality_donut: categoria com cardinalidade 2-8 + 1 metrica -> donut chart",
                                recommended_aggregation="sum",
                            ),
                        )

        # HEURISTICA 4: 2 metricas = scatter
        if len(numeric_cols) >= 2:
            for i, num1 in enumerate(numeric_cols[:3]):
                for num2 in numeric_cols[i + 1:4]:
                    self._append_suggestion(
                        suggestions,
                        seen_suggestions,
                        ChartSuggestion(
                            chart_type="scatter",
                            title=f"Correlacao: {num1.name} vs {num2.name}",
                            description=f"Relacao entre {num1.name} e {num2.name}",
                            dimension_columns=[],
                            metric_columns=[num1.name, num2.name],
                            confidence_score=0.80,
                            heuristic_rule="two_metrics_scatter: 2 metricas numericas -> scatter plot",
                            recommended_aggregation="avg",
                        ),
                    )

        # HEURISTICA 5: Multiplas metricas por categoria = barras agrupadas
        if categorical_cols and len(numeric_cols) >= 2:
            for cat in categorical_cols[:2]:
                if cat.cardinality <= 20:
                    self._append_suggestion(
                        suggestions,
                        seen_suggestions,
                        ChartSuggestion(
                            chart_type="bar",
                            title=f"Metricas por {cat.name}",
                            description=f"Comparacao multipla por {cat.name}",
                            dimension_columns=[cat.name],
                            metric_columns=[n.name for n in numeric_cols[:4]],
                            confidence_score=0.85,
                            heuristic_rule="multi_metric_grouped_bar: 1 categoria + multiplas metricas -> grouped bar",
                            recommended_aggregation="sum",
                        ),
                    )

        # HEURISTICA 6: Hierarquia = treemap
        if len(categorical_cols) >= 2:
            hierarchy = [c.name for c in categorical_cols[:3]]
            if numeric_cols:
                self._append_suggestion(
                    suggestions,
                    seen_suggestions,
                    ChartSuggestion(
                        chart_type="treemap",
                        title="Hierarquia de categorias",
                        description=f"Visualizacao hierarquica: {' > '.join(hierarchy)}",
                        dimension_columns=hierarchy,
                        metric_columns=[numeric_cols[0].name],
                        confidence_score=0.75,
                        heuristic_rule="hierarchy_treemap: multiplas colunas categoricas -> treemap",
                        recommended_aggregation="sum",
                    ),
                )

        # HEURISTICA 7: KPI unico = radial bar ou card
        if numeric_cols:
            for num in numeric_cols[:2]:
                self._append_suggestion(
                    suggestions,
                    seen_suggestions,
                    ChartSuggestion(
                        chart_type="radial_bar",
                        title=f"KPI: {num.name}",
                        description=f"Indicador unico: {num.name}",
                        dimension_columns=[],
                        metric_columns=[num.name],
                        confidence_score=0.70,
                        heuristic_rule="single_kpi_radial: 1 metrica isolada -> radial bar / KPI card",
                        recommended_aggregation="sum",
                    ),
                )
                self._append_suggestion(
                    suggestions,
                    seen_suggestions,
                    ChartSuggestion(
                        chart_type="kpi",
                        title=f"KPI: {num.name}",
                        description=f"Indicador unico: {num.name}",
                        dimension_columns=[],
                        metric_columns=[num.name],
                        confidence_score=0.70,
                        heuristic_rule="single_kpi_card: 1 metrica isolada -> KPI card",
                        recommended_aggregation="sum",
                    ),
                )

        # HEURISTICA 8: Radar para multiplas metricas
        if len(numeric_cols) >= 3 and categorical_cols:
            for cat in categorical_cols[:1]:
                if cat.cardinality <= 10:
                    self._append_suggestion(
                        suggestions,
                        seen_suggestions,
                        ChartSuggestion(
                            chart_type="radar",
                            title=f"Perfil por {cat.name}",
                            description=f"Comparacao multidimensional por {cat.name}",
                            dimension_columns=[cat.name],
                            metric_columns=[n.name for n in numeric_cols[:6]],
                            confidence_score=0.72,
                            heuristic_rule="multi_metric_radar: 1 categoria + 3+ metricas -> radar chart",
                            recommended_aggregation="avg",
                        ),
                    )

        suggestions.sort(key=lambda x: x.confidence_score, reverse=True)
        return suggestions[:12]

    def _build_financial_suggestions(self, columns: list[ColumnSchema]) -> list[ChartSuggestion]:
        """Cria sugestoes prioritarias para cenarios financeiros mensais, anuais e diarios."""
        if not columns:
            return []

        numeric_cols = [c for c in columns if self._is_numeric_metric_candidate(c)]
        dimension_cols = [
            c for c in columns
            if c.detected_type in ("categorical", "string", "boolean", "date", "datetime") and not self._is_generic_column_name(c.name)
        ]

        period_columns = {
            "monthly": self._find_first_matching_column(columns, ("mes", "month", "competencia", "periodo", "periodo_ref", "referencia")),
            "yearly": self._find_first_matching_column(columns, ("ano", "year", "anual", "exercicio")),
            "daily": self._find_first_matching_column(columns, ("dia", "day", "data", "date", "dt_", "data_ref")),
        }
        revenue_col = self._find_first_matching_metric(
            numeric_cols,
            (
                "receita", "faturamento", "entrada", "ganho", "venda", "receb", "credito",
                "pix", "boleto", "cartao", "ted", "doc", "transferencia", "dinheiro",
            ),
        )
        expense_col = self._find_first_matching_metric(
            numeric_cols,
            (
                "despesa", "custo", "gasto", "saida", "pagamento", "pag", "pagamentos", "debito",
                "boleto", "pix", "cartao", "ted", "doc", "transferencia", "dinheiro",
            ),
        )
        balance_col = self._find_first_matching_metric(numeric_cols, ("saldo", "caixa", "disponivel", "acumulado"))
        result_col = self._find_first_matching_metric(numeric_cols, ("superavit", "deficit", "resultado", "lucro", "prejuizo"))
        realized_col = self._find_first_matching_metric(
            numeric_cols,
            ("realizada", "realizado", "recebida", "recebido", "recebimentos"),
        )
        billed_col = self._find_first_matching_metric(
            numeric_cols,
            ("faturada", "faturado", "emitida", "emitido", "projetada", "prevista"),
        )
        expense_breakdown_col = self._find_best_dimension(
            dimension_cols,
            (
                "despesa", "categoria", "tipo", "conta", "grupo", "centro", "natureza",
                "pagamento", "forma", "meio", "canal", "cartao", "ted", "doc", "transferencia", "dinheiro",
            ),
        )
        revenue_breakdown_col = self._find_best_dimension(
            dimension_cols,
            (
                "receita", "status", "origem", "tipo", "categoria", "canal", "pagamento",
                "forma", "meio", "pix", "boleto", "cartao", "ted", "doc", "transferencia", "dinheiro",
            ),
        )

        suggestions: list[ChartSuggestion] = []

        def add_period_suggestion(
            period_key: str,
            period_label: str,
            metric: ColumnSchema | None,
            metric_label: str,
            chart_type: str,
            score: float,
            rule: str,
        ) -> None:
            dimension = period_columns.get(period_key)
            if not dimension or not metric:
                return
            suggestions.append(
                ChartSuggestion(
                    chart_type=chart_type,
                    title=f"{metric_label} {period_label}",
                    description=f"Acompanhamento {period_label} de {metric.name} usando a coluna {dimension.name}.",
                    dimension_columns=[dimension.name],
                    metric_columns=[metric.name],
                    confidence_score=score,
                    heuristic_rule=rule,
                    recommended_aggregation="sum",
                )
            )

        add_period_suggestion("monthly", "mensais", revenue_col, "Receitas", "line", 0.995, "finance_monthly_revenue: coluna de periodo mensal + metrica de receita -> line chart")
        add_period_suggestion("monthly", "mensais", expense_col, "Despesas", "bar", 0.992, "finance_monthly_expense: coluna de periodo mensal + metrica de despesa -> bar chart")
        add_period_suggestion("monthly", "mensais", result_col or balance_col, "Deficit / superavit", "bar", 0.989, "finance_monthly_result: coluna de periodo mensal + metrica de resultado/saldo -> bar chart")
        add_period_suggestion("yearly", "anuais", revenue_col, "Receitas", "bar", 0.991, "finance_yearly_revenue: coluna anual + metrica de receita -> bar chart")
        add_period_suggestion("yearly", "anuais", expense_col, "Despesas", "bar", 0.988, "finance_yearly_expense: coluna anual + metrica de despesa -> bar chart")
        add_period_suggestion("yearly", "anuais", result_col or balance_col, "Deficit / superavit", "bar", 0.985, "finance_yearly_result: coluna anual + metrica de resultado/saldo -> bar chart")
        add_period_suggestion("yearly", "anuais", balance_col, "Saldos", "line", 0.983, "finance_yearly_balance: coluna anual + metrica de saldo -> line chart")
        add_period_suggestion("daily", "diarios", revenue_col or result_col or expense_col, "Evolucao", "line", 0.98, "finance_daily_overview: coluna diaria + metrica financeira principal -> line chart")

        if revenue_breakdown_col and (realized_col or billed_col or revenue_col):
            metric = realized_col or billed_col or revenue_col
            suggestions.append(
                ChartSuggestion(
                    chart_type="donut",
                    title="Divisao das receitas",
                    description=f"Composicao das receitas por {revenue_breakdown_col.name}, priorizando realizadas e faturadas quando existirem.",
                    dimension_columns=[revenue_breakdown_col.name],
                    metric_columns=[metric.name],
                    confidence_score=0.994,
                    heuristic_rule="finance_revenue_split: categoria de receita/status + metrica de receita -> donut chart",
                    recommended_aggregation="sum",
                )
            )

        if expense_breakdown_col and (expense_col or result_col):
            metric = expense_col or result_col
            suggestions.append(
                ChartSuggestion(
                    chart_type="pie",
                    title="Divisao das despesas",
                    description=(
                        f"Distribuicao das despesas por {expense_breakdown_col.name}, "
                        "com destaque para impostos, pecas, folha, vendas e sede quando presentes."
                    ),
                    dimension_columns=[expense_breakdown_col.name],
                    metric_columns=[metric.name],
                    confidence_score=0.993,
                    heuristic_rule="finance_expense_split: categoria de despesa + metrica financeira -> pie chart",
                    recommended_aggregation="sum",
                )
            )

        # Fallback para planilhas de fluxo no formato "wide" (meses em colunas).
        if not period_columns.get("monthly"):
            wide_month_columns = [column for column in columns if self._is_month_column_name(column.name)]
            description_column = self._find_best_dimension(
                dimension_cols,
                ("descricao", "descrição", "conta", "historico", "histórico", "detalhamento", "referencia", "referência"),
            )

            if description_column and len(wide_month_columns) >= 3:
                suggestions.extend(
                    [
                        ChartSuggestion(
                            chart_type="line",
                            title="Receitas mensais",
                            description="Serie mensal de receitas consolidada a partir do fluxo financeiro por descricao.",
                            dimension_columns=["__month"],
                            metric_columns=["__finance_receitas"],
                            confidence_score=0.997,
                            heuristic_rule=(
                                "finance_wide_monthly_revenue: colunas mensais + descricao -> serie mensal de receitas"
                            ),
                            recommended_aggregation="sum",
                        ),
                        ChartSuggestion(
                            chart_type="bar",
                            title="Despesas mensais",
                            description="Serie mensal de despesas consolidada a partir do fluxo financeiro por descricao.",
                            dimension_columns=["__month"],
                            metric_columns=["__finance_despesas"],
                            confidence_score=0.996,
                            heuristic_rule=(
                                "finance_wide_monthly_expense: colunas mensais + descricao -> serie mensal de despesas"
                            ),
                            recommended_aggregation="sum",
                        ),
                        ChartSuggestion(
                            chart_type="bar",
                            title="Deficit / superavit mensais",
                            description="Resultado mensal (receitas - despesas) consolidado da planilha de fluxo.",
                            dimension_columns=["__month"],
                            metric_columns=["__finance_resultado"],
                            confidence_score=0.995,
                            heuristic_rule=(
                                "finance_wide_monthly_result: colunas mensais + descricao -> resultado mensal"
                            ),
                            recommended_aggregation="sum",
                        ),
                        ChartSuggestion(
                            chart_type="line",
                            title="Saldos anuais",
                            description="Curva de saldos consolidada pelo fluxo para leitura anual do periodo disponivel.",
                            dimension_columns=["__month"],
                            metric_columns=["__finance_saldos"],
                            confidence_score=0.994,
                            heuristic_rule=(
                                "finance_wide_balance: colunas mensais + descricao -> curva de saldos"
                            ),
                            recommended_aggregation="sum",
                        ),
                        ChartSuggestion(
                            chart_type="donut",
                            title="Divisao das receitas",
                            description="Composicao de receitas realizadas e faturadas inferida do detalhamento de receita.",
                            dimension_columns=["__finance_receita_status"],
                            metric_columns=["__finance_receitas_split"],
                            confidence_score=0.993,
                            heuristic_rule=(
                                "finance_wide_revenue_split: fluxo wide -> divisao receitas realizadas/faturadas"
                            ),
                            recommended_aggregation="sum",
                        ),
                        ChartSuggestion(
                            chart_type="pie",
                            title="Divisao das despesas",
                            description="Composicao de despesas por categorias principais (impostos, pecas, folha, vendas, sede).",
                            dimension_columns=["__finance_despesa_categoria"],
                            metric_columns=["__finance_despesas_split"],
                            confidence_score=0.992,
                            heuristic_rule=(
                                "finance_wide_expense_split: fluxo wide -> divisao despesas por categoria"
                            ),
                            recommended_aggregation="sum",
                        ),
                    ]
                )

        return suggestions

    def _append_suggestion(
        self,
        suggestions: list[ChartSuggestion],
        seen_suggestions: set[tuple[str, tuple[str, ...], tuple[str, ...], str]],
        suggestion: ChartSuggestion,
    ) -> None:
        key = (
            suggestion.chart_type,
            tuple(suggestion.dimension_columns),
            tuple(suggestion.metric_columns),
            suggestion.recommended_aggregation,
        )
        if key in seen_suggestions:
            return
        seen_suggestions.add(key)
        suggestions.append(suggestion)

    def _normalize_name(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")

    def _find_first_matching_column(self, columns: list[ColumnSchema], keywords: tuple[str, ...]) -> ColumnSchema | None:
        matches = self._find_matching_columns(columns, keywords)
        return matches[0] if matches else None

    def _find_first_matching_metric(self, columns: list[ColumnSchema], keywords: tuple[str, ...]) -> ColumnSchema | None:
        matches = self._find_matching_columns(columns, keywords)
        return matches[0] if matches else None

    def _find_best_dimension(self, columns: list[ColumnSchema], keywords: tuple[str, ...]) -> ColumnSchema | None:
        matches = self._find_matching_columns(columns, keywords)
        if not matches:
            return None
        return sorted(matches, key=lambda col: (col.cardinality == 0, col.cardinality or 10**6))[0]

    def _find_matching_columns(self, columns: list[ColumnSchema], keywords: tuple[str, ...]) -> list[ColumnSchema]:
        normalized_keywords = tuple(self._normalize_name(keyword) for keyword in keywords)
        ranked: list[tuple[int, ColumnSchema]] = []
        for column in columns:
            normalized = self._normalize_name(column.name)
            score = sum(1 for keyword in normalized_keywords if keyword and keyword in normalized)
            if score:
                ranked.append((score, column))
        ranked.sort(key=lambda item: (-item[0], item[1].index))
        return [column for _, column in ranked]

    def get_chart_data(self, file_uuid: str, sheet_name: str, chart_type: str,
                       dimension_columns: list[str], metric_columns: list[str],
                       aggregation: str = "sum", filters: Optional[dict] = None,
                       limit: int = 1000) -> dict[str, Any]:
        """Gera dados agregados para um grafico especifico."""

        db_path = self.storage_path / file_uuid / "analytics.db"
        if not db_path.exists():
            raise FileNotFoundError(f"Banco de dados para {file_uuid} nao encontrado.")

        conn = duckdb.connect(str(db_path), read_only=True)
        table_name = self._sanitize_table_name(sheet_name)

        if metric_columns and metric_columns[0].startswith("__finance_"):
            data = self._get_wide_finance_chart_data(
                conn,
                table_name,
                metric_columns[0],
                dimension_columns,
                file_uuid,
                sheet_name,
            )
            conn.close()
            return {
                "chart_type": chart_type,
                "dimensions": dimension_columns,
                "metrics": metric_columns,
                "data": data,
                "total_rows": len(data),
                "applied_filters": filters or {},
                "generated_at": datetime.utcnow().isoformat(),
            }

        dims = [f'"{d}"' for d in dimension_columns]
        mets = [f'"{m}"' for m in metric_columns]

        # Scatter plot nao tem agregacao, retorna valores brutos
        if chart_type == "scatter" and len(mets) >= 2:
            select_cols = mets
            group_by = ""
        else:
            agg_funcs = {"sum": "SUM", "avg": "AVG", "count": "COUNT",
                         "min": "MIN", "max": "MAX", "distinct": "COUNT(DISTINCT"}
            agg = agg_funcs.get(aggregation, "SUM")

            agg_metrics = []
            for m_col in metric_columns:
                if aggregation == "distinct":
                    agg_metrics.append(f'{agg}("{m_col}")) AS "{m_col}"')
                elif aggregation == "count":
                    agg_metrics.append(f'{agg}("{m_col}") AS "{m_col}"')
                else:
                    # DuckDB pode ter colunas numericas como VARCHAR (ex: CSV).
                    # TRY_CAST evita falha e torna agregacao robusta.
                    agg_metrics.append(f'{agg}(TRY_CAST("{m_col}" AS DOUBLE)) AS "{m_col}"')

            select_cols = dims + agg_metrics
            group_by = f"GROUP BY {', '.join(dims)}" if dims else ""

        where_clause = ""
        if filters:
            conditions = []

            def _escape(value: Any) -> str:
                return str(value).replace("'", "''")

            def _to_number(value: Any) -> float:
                if isinstance(value, str):
                    value = value.replace(",", ".")
                return float(value)

            for col, val in filters.items():
                # Sanitiza nome da coluna para evitar injecao
                safe_col = f'"{col.replace("\"", "")}"'

                if isinstance(val, list):
                    placeholders = ", ".join(f"'{_escape(v)}'" for v in val)
                    conditions.append(f'{safe_col} IN ({placeholders})')
                elif isinstance(val, dict) and "operator" in val:
                    op = str(val.get("operator", "equals"))
                    op_value = val.get("value")
                    op_value_to = val.get("value_to")

                    if op == "equals":
                        conditions.append(f"{safe_col} = '{_escape(op_value)}'")
                    elif op == "not_equals":
                        conditions.append(f"{safe_col} <> '{_escape(op_value)}'")
                    elif op == "contains":
                        conditions.append(f"CAST({safe_col} AS VARCHAR) LIKE '%{_escape(op_value)}%'")
                    elif op == "not_contains":
                        conditions.append(f"CAST({safe_col} AS VARCHAR) NOT LIKE '%{_escape(op_value)}%'")
                    elif op == "starts_with":
                        conditions.append(f"CAST({safe_col} AS VARCHAR) LIKE '{_escape(op_value)}%'")
                    elif op == "ends_with":
                        conditions.append(f"CAST({safe_col} AS VARCHAR) LIKE '%{_escape(op_value)}'")
                    elif op == "greater_than":
                        conditions.append(f"TRY_CAST({safe_col} AS DOUBLE) > {_to_number(op_value)}")
                    elif op == "less_than":
                        conditions.append(f"TRY_CAST({safe_col} AS DOUBLE) < {_to_number(op_value)}")
                    elif op == "between":
                        conditions.append(
                            f"TRY_CAST({safe_col} AS DOUBLE) BETWEEN {_to_number(op_value)} AND {_to_number(op_value_to)}"
                        )
                    elif op == "in":
                        values = op_value if isinstance(op_value, list) else [v.strip() for v in str(op_value).split(",") if v.strip()]
                        placeholders = ", ".join(f"'{_escape(v)}'" for v in values)
                        conditions.append(f"CAST({safe_col} AS VARCHAR) IN ({placeholders})")
                    elif op == "not_in":
                        values = op_value if isinstance(op_value, list) else [v.strip() for v in str(op_value).split(",") if v.strip()]
                        placeholders = ", ".join(f"'{_escape(v)}'" for v in values)
                        conditions.append(f"CAST({safe_col} AS VARCHAR) NOT IN ({placeholders})")
                    elif op == "is_null":
                        conditions.append(f"{safe_col} IS NULL")
                    elif op == "is_not_null":
                        conditions.append(f"{safe_col} IS NOT NULL")
                elif isinstance(val, dict) and "min" in val and "max" in val:
                    conditions.append(f'TRY_CAST({safe_col} AS DOUBLE) BETWEEN {_to_number(val["min"])} AND {_to_number(val["max"])}')
                else:
                    conditions.append(f"{safe_col} = '{_escape(val)}'")
            if conditions:
                where_clause = "WHERE " + " AND ".join(conditions)

        order_by = f"ORDER BY {dims[0]}" if dims else ""
        limit_clause = f"LIMIT {min(limit, 10000)}"  # Hard limit

        query = f"SELECT {', '.join(select_cols)} FROM {table_name} {where_clause} {group_by} {order_by} {limit_clause}"

        try:
            logger.debug(f"Executando query de chart data: {query}")
            result_df = conn.execute(query).fetchdf()
            data = result_df.to_dict("records")
            total_rows = len(data)
        except Exception as e:
            logger.error(f"Erro na query de chart data: {e} | Query: {query}")
            data = []
            total_rows = 0
        finally:
            conn.close()

        return {
            "chart_type": chart_type,
            "dimensions": dimension_columns,
            "metrics": metric_columns,
            "data": data,
            "total_rows": total_rows,
            "applied_filters": filters or {},
            "generated_at": datetime.utcnow().isoformat(),
        }

    def get_table_data(self, file_uuid: str, sheet_name: str, page: int = 1,
                       page_size: int = 100, sort_by: Optional[str] = None,
                       sort_direction: str = "asc", filters: Optional[list] = None,
                       visible_columns: Optional[list[str]] = None) -> dict[str, Any]:
        """Retorna dados paginados para tabela."""

        db_path = self.storage_path / file_uuid / "analytics.db"
        conn = duckdb.connect(str(db_path), read_only=True)
        table_name = self._sanitize_table_name(sheet_name)

        offset = (page - 1) * page_size

        if visible_columns:
            cols = ", ".join(f'"{c}"' for c in visible_columns)
        else:
            cols = "*"

        where_clause = ""
        if filters:
            conditions = []

            def _escape(value: Any) -> str:
                return str(value).replace("'", "''")

            def _to_number(value: Any) -> float:
                if isinstance(value, str):
                    value = value.replace(",", ".")
                return float(value)

            for f in filters:
                col = f.get("column")
                op = f.get("operator")
                val = f.get("value")
                safe_col = f'"{str(col).replace("\"", "")}"'

                if op == "equals":
                    conditions.append(f"{safe_col} = '{_escape(val)}'")
                elif op == "not_equals":
                    conditions.append(f"{safe_col} <> '{_escape(val)}'")
                elif op == "contains":
                    conditions.append(f"CAST({safe_col} AS VARCHAR) LIKE '%{_escape(val)}%'")
                elif op == "not_contains":
                    conditions.append(f"CAST({safe_col} AS VARCHAR) NOT LIKE '%{_escape(val)}%'")
                elif op == "starts_with":
                    conditions.append(f"CAST({safe_col} AS VARCHAR) LIKE '{_escape(val)}%'")
                elif op == "ends_with":
                    conditions.append(f"CAST({safe_col} AS VARCHAR) LIKE '%{_escape(val)}'")
                elif op == "greater_than":
                    conditions.append(f"TRY_CAST({safe_col} AS DOUBLE) > {_to_number(val)}")
                elif op == "less_than":
                    conditions.append(f"TRY_CAST({safe_col} AS DOUBLE) < {_to_number(val)}")
                elif op == "between":
                    val_to = f.get("value_to")
                    conditions.append(f"TRY_CAST({safe_col} AS DOUBLE) BETWEEN {_to_number(val)} AND {_to_number(val_to)}")
                elif op == "in":
                    values = val if isinstance(val, list) else [v.strip() for v in str(val).split(",") if v.strip()]
                    placeholders = ", ".join(f"'{_escape(v)}'" for v in values)
                    conditions.append(f"CAST({safe_col} AS VARCHAR) IN ({placeholders})")
                elif op == "not_in":
                    values = val if isinstance(val, list) else [v.strip() for v in str(val).split(",") if v.strip()]
                    placeholders = ", ".join(f"'{_escape(v)}'" for v in values)
                    conditions.append(f"CAST({safe_col} AS VARCHAR) NOT IN ({placeholders})")
                elif op == "is_null":
                    conditions.append(f"{safe_col} IS NULL")
                elif op == "is_not_null":
                    conditions.append(f"{safe_col} IS NOT NULL")

            if conditions:
                where_clause = "WHERE " + " AND ".join(conditions)

        order_clause = ""
        if sort_by:
            direction = "DESC" if sort_direction == "desc" else "ASC"
            order_clause = f'ORDER BY "{sort_by}" {direction}'

        query = f"SELECT {cols} FROM {table_name} {where_clause} {order_clause} LIMIT {page_size} OFFSET {offset}"
        count_query = f"SELECT COUNT(*) as total FROM {table_name} {where_clause}"

        try:
            result_df = conn.execute(query).fetchdf()
            count_result = conn.execute(count_query).fetchone()
            total_rows = count_result[0] if count_result else 0
            data = result_df.to_dict("records")
        except Exception as e:
            logger.error(f"Erro na query de tabela: {e}")
            data = []
            total_rows = 0
        finally:
            conn.close()

        total_pages = (total_rows + page_size - 1) // page_size

        return {
            "data": data,
            "page": page,
            "page_size": page_size,
            "total_rows": total_rows,
            "total_pages": total_pages,
            "applied_filters": filters or [],
        }

    def _extract_column_info(self, conn, table_name: str, schema_df) -> list[ColumnSchema]:
        """Extrai informacoes detalhadas das colunas via DuckDB."""
        columns = []

        for idx, row in schema_df.iterrows():
            col_name = row["column_name"]
            col_type = row["column_type"]
            quoted_col = f'"{col_name}"'

            try:
                stats_query = (
                    f"SELECT "
                    f"COUNT(*) as total, "
                    f"COUNT(*) FILTER (WHERE {quoted_col} IS NULL) as null_count, "
                    f"COUNT(DISTINCT {quoted_col}) as unique_count "
                    f"FROM {table_name}"
                )
                total, null_count, unique_count = conn.execute(stats_query).fetchone()
                cardinality = unique_count
                samples_df = conn.execute(
                    f"SELECT {quoted_col} FROM {table_name} "
                    f"WHERE {quoted_col} IS NOT NULL "
                    f"AND TRIM(CAST({quoted_col} AS VARCHAR)) <> '' "
                    f"LIMIT 50"
                ).fetchdf()
                samples = [str(v) for v in samples_df[col_name].tolist()]
            except Exception:
                total = 0
                unique_count = 0
                null_count = 0
                cardinality = 0
                samples = []

            detected_type = self._map_duckdb_type(col_type)
            if detected_type == "string":
                detected_type = self._infer_string_type(samples)

            col_schema = ColumnSchema(
                name=col_name,
                index=idx,
                detected_type=detected_type,
                null_count=int(null_count),
                unique_count=int(unique_count),
                cardinality=int(cardinality),
                sample_values=samples[:10],
            )
            columns.append(col_schema)

        return columns

    def _map_duckdb_type(self, duckdb_type: str) -> ColumnType:
        """Mapeia tipo DuckDB para tipo interno."""
        dt = duckdb_type.upper()

        if "VARCHAR" in dt or "TEXT" in dt or "CHAR" in dt:
            return "string"
        elif "INTEGER" in dt or "BIGINT" in dt or "SMALLINT" in dt or "TINYINT" in dt:
            return "integer"
        elif "DOUBLE" in dt or "FLOAT" in dt or "REAL" in dt:
            return "float"
        elif "DECIMAL" in dt or "NUMERIC" in dt:
            return "decimal"
        elif "BOOLEAN" in dt:
            return "boolean"
        elif "DATE" in dt and "TIME" not in dt:
            return "date"
        elif "TIMESTAMP" in dt or "DATETIME" in dt:
            return "datetime"
        else:
            return "mixed"

    def _infer_string_type(self, samples: list[str]) -> ColumnType:
        """Infere tipos especiais em colunas string de forma deterministica."""
        if not samples:
            return "empty"

        filtered_samples = [s.strip() for s in samples if str(s).strip()]
        filtered_samples = [s for s in filtered_samples if not s.startswith("=")]

        if not filtered_samples:
            return "empty"

        lowered = [s.lower() for s in filtered_samples]
        if not lowered:
            return "empty"

        if all(_PERCENT_RE.match(s) for s in filtered_samples):
            return "percentage"
        if all(_CURRENCY_RE.match(s) for s in filtered_samples):
            return "currency"

        bool_values = {"true", "false", "yes", "no", "sim", "nao", "1", "0"}
        if all(v in bool_values for v in lowered):
            return "boolean"

        date_hits = sum(1 for s in filtered_samples if self._looks_like_date(s))
        if date_hits / len(filtered_samples) >= 0.85:
            return "date"

        int_hits = sum(1 for s in filtered_samples if _INT_RE.match(s))
        float_hits = sum(1 for s in filtered_samples if _FLOAT_RE.match(s))
        numeric_hits = int_hits + float_hits
        if int_hits == len(filtered_samples):
            return "integer"
        if numeric_hits == len(filtered_samples) and float_hits > 0:
            return "float"
        if numeric_hits / len(filtered_samples) >= 0.8:
            return "float"

        uniq_ratio = len(set(lowered)) / len(lowered)
        if len(set(lowered)) <= 50 and uniq_ratio <= 0.2:
            return "categorical"

        return "string"

    def _is_numeric_metric_candidate(self, column: ColumnSchema) -> bool:
        if self._is_generic_column_name(column.name):
            return False

        if column.detected_type in ("integer", "float", "decimal", "currency", "percentage"):
            return True

        if column.detected_type not in ("string", "mixed"):
            return False

        samples = [value for value in column.sample_values if str(value).strip()]
        if not samples:
            return False

        parsed_hits = sum(1 for value in samples if self._parse_numeric_value(value) is not None)
        ratio = parsed_hits / len(samples)

        if self._is_month_column_name(column.name):
            return ratio >= 0.3

        return ratio >= 0.7

    def _is_generic_column_name(self, name: str) -> bool:
        normalized = self._normalize_name(name)
        if normalized == "":
            return True
        if normalized.startswith("unnamed"):
            return True
        if re.fullmatch(r"col_\d+", normalized):
            return True
        if re.fullmatch(r"\d+", normalized):
            return True
        return False

    def _is_month_column_name(self, name: str) -> bool:
        normalized = self._normalize_name(self._normalize_label(name))
        month_tokens = {
            "jan", "janeiro", "feb", "fev", "fevereiro", "mar", "marco", "abril", "abr",
            "mai", "maio", "jun", "junho", "jul", "julho", "ago", "agosto",
            "set", "setembro", "out", "outubro", "nov", "novembro", "dez", "dezembro",
        }
        return normalized in month_tokens

    def _parse_numeric_value(self, value: Any) -> float | None:
        text = str(value).strip()
        if not text or text.startswith("="):
            return None

        if text in {"-", "--", ".", ""}:
            return 0.0

        negative = False
        if text.startswith("(") and text.endswith(")"):
            negative = True
            text = text[1:-1].strip()

        cleaned = text.replace("R$", "").replace("$", "").replace("€", "").replace("£", "").replace("¥", "")
        cleaned = cleaned.replace("%", "").replace(" ", "")

        # "1.234,56" -> "1234.56" e "1,234.56" -> "1234.56"
        if "," in cleaned and "." in cleaned:
            if cleaned.rfind(",") > cleaned.rfind("."):
                cleaned = cleaned.replace(".", "").replace(",", ".")
            else:
                cleaned = cleaned.replace(",", "")
        else:
            if "." in cleaned and re.fullmatch(r"\d{1,3}(?:\.\d{3})+", cleaned):
                cleaned = cleaned.replace(".", "")
            cleaned = cleaned.replace(",", ".")

        try:
            parsed = float(cleaned)
            return -parsed if negative else parsed
        except ValueError:
            return None

    def _looks_like_date(self, value: str) -> bool:
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
        return not pd.isna(parsed)

    def _sanitize_table_name(self, name: str) -> str:
        """Sanitiza nome para SQL."""
        sanitized = "".join(c if c.isalnum() else "_" for c in name)
        return f"sheet_{sanitized}"[:60]

    def _get_wide_finance_chart_data(
        self,
        conn,
        table_name: str,
        metric_tag: str,
        dimension_columns: list[str],
        file_uuid: str,
        sheet_name: str,
    ) -> list[dict[str, Any]]:
        workbook_data = self._extract_wide_finance_from_workbook(file_uuid, sheet_name, metric_tag)
        if workbook_data:
            return workbook_data

        frame = conn.execute(f"SELECT * FROM {table_name}").fetchdf()
        if frame.empty:
            return []

        description_col = self._find_description_column(frame.columns)
        month_columns = [column for column in frame.columns if self._is_month_column_name(column)]
        if not description_col or not month_columns:
            return []

        descriptions = frame[description_col].fillna("").astype(str)

        revenue_mask = descriptions.str.contains(r"receita|faturamento|venda|entrada|receb", case=False, regex=True)
        expense_mask = descriptions.str.contains(r"despesa|custo|gasto|imposto|folha|sede|pe[cç]a|comiss", case=False, regex=True)
        balance_mask = descriptions.str.contains(r"saldo|caixa", case=False, regex=True)
        result_mask = descriptions.str.contains(r"resultado|super[áa]vit|def[íi]cit|lucro|preju[íi]zo", case=False, regex=True)

        def month_sum(mask) -> dict[str, float]:
            totals: dict[str, float] = OrderedDict()
            selected = frame[mask]
            for column in month_columns:
                values = selected[column] if column in selected else []
                totals[column] = float(sum(self._parse_numeric_value(value) or 0.0 for value in values))
            return totals

        revenue_by_month = month_sum(revenue_mask)
        expense_by_month = month_sum(expense_mask)
        balance_by_month = month_sum(balance_mask)
        result_by_month = month_sum(result_mask)

        if metric_tag == "__finance_receitas":
            return [{"__month": month, metric_tag: value} for month, value in revenue_by_month.items()]

        if metric_tag == "__finance_despesas":
            return [{"__month": month, metric_tag: value} for month, value in expense_by_month.items()]

        if metric_tag == "__finance_saldos":
            payload = balance_by_month
            if not any(abs(value) > 0 for value in payload.values()):
                payload = {
                    month: revenue_by_month.get(month, 0.0) - expense_by_month.get(month, 0.0)
                    for month in month_columns
                }
            return [{"__month": month, metric_tag: value} for month, value in payload.items()]

        if metric_tag == "__finance_resultado":
            payload = result_by_month
            if not any(abs(value) > 0 for value in payload.values()):
                payload = {
                    month: revenue_by_month.get(month, 0.0) - expense_by_month.get(month, 0.0)
                    for month in month_columns
                }
            return [{"__month": month, metric_tag: value} for month, value in payload.items()]

        if metric_tag == "__finance_receitas_split":
            labels = descriptions
            receita_rows = frame[revenue_mask]
            realized_mask = labels.str.contains(r"pix|boleto|cart[aã]o|cartoes|cartões|dinheiro|recebida|realizada", case=False, regex=True)
            billed_mask = labels.str.contains(r"faturad|emitid|previst", case=False, regex=True)

            realized_total = 0.0
            billed_total = 0.0
            for column in month_columns:
                realized_total += float(sum(self._parse_numeric_value(v) or 0.0 for v in frame[realized_mask][column]))
                billed_total += float(sum(self._parse_numeric_value(v) or 0.0 for v in frame[billed_mask][column]))

            if billed_total == 0.0:
                # Mantem o shape solicitado mesmo quando o arquivo nao separa faturadas explicitamente.
                billed_total = max(0.0, float(sum(self._parse_numeric_value(v) or 0.0 for col in month_columns for v in receita_rows[col])) - realized_total)

            return [
                {"__finance_receita_status": "Realizadas", metric_tag: realized_total},
                {"__finance_receita_status": "Faturada", metric_tag: billed_total},
            ]

        if metric_tag == "__finance_despesas_split":
            categories = {
                "Impostos": r"imposto|taxa|tribut",
                "Pecas nac.": r"pe[cç]a.*nac|nacional",
                "Pecas import.": r"pe[cç]a.*import|importad",
                "Folha": r"folha|salario|sal[áa]rio|prolabore|pro labore|13o|13º",
                "Vendas": r"comiss|venda|comercial",
                "Sede": r"sede|administr|escritorio|escrit[óo]rio|aluguel",
            }

            results: list[dict[str, Any]] = []
            for category, pattern in categories.items():
                mask = descriptions.str.contains(pattern, case=False, regex=True)
                total = 0.0
                for column in month_columns:
                    total += float(sum(self._parse_numeric_value(v) or 0.0 for v in frame[mask][column]))
                results.append({"__finance_despesa_categoria": category, metric_tag: total})

            other_mask = expense_mask.copy()
            for pattern in categories.values():
                other_mask = other_mask & ~descriptions.str.contains(pattern, case=False, regex=True)
            other_total = 0.0
            for column in month_columns:
                other_total += float(sum(self._parse_numeric_value(v) or 0.0 for v in frame[other_mask][column]))
            if other_total:
                results.append({"__finance_despesa_categoria": "Outras", metric_tag: other_total})

            return results

        return []

    def _find_description_column(self, columns: list[str]) -> str | None:
        normalized_columns = [(column, self._normalize_name(column)) for column in columns]
        for original, normalized in normalized_columns:
            if any(token in normalized for token in ("descricao", "descri", "conta", "historico", "referencia", "detalhamento")):
                return original
        return None

    def _normalize_label(self, value: Any) -> str:
        text = str(value or "").strip().lower()
        normalized = unicodedata.normalize("NFD", text)
        normalized = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized

    def _month_sort_key(self, month_label: str) -> int:
        normalized = self._normalize_name(self._normalize_label(month_label))
        order = {
            "jan": 1,
            "janeiro": 1,
            "fev": 2,
            "fevereiro": 2,
            "mar": 3,
            "marco": 3,
            "abr": 4,
            "abril": 4,
            "mai": 5,
            "maio": 5,
            "jun": 6,
            "junho": 6,
            "jul": 7,
            "julho": 7,
            "ago": 8,
            "agosto": 8,
            "set": 9,
            "setembro": 9,
            "out": 10,
            "outubro": 10,
            "nov": 11,
            "novembro": 11,
            "dez": 12,
            "dezembro": 12,
        }
        return order.get(normalized, 99)

    def _resolve_original_workbook_path(self, file_uuid: str) -> Path | None:
        base = self.storage_path / file_uuid
        if not base.exists():
            return None

        for extension in (".xlsx", ".xlsm", ".xls"):
            candidate = base / f"original{extension}"
            if candidate.exists():
                return candidate

        matches = sorted(base.glob("original.*"))
        return matches[0] if matches else None

    def _find_description_col_index(self, worksheet) -> int | None:
        max_rows = min(60, worksheet.max_row)
        max_cols = min(25, worksheet.max_column)

        for row in range(1, max_rows + 1):
            for col in range(1, max_cols + 1):
                value = worksheet.cell(row=row, column=col).value
                normalized = self._normalize_label(value)
                if normalized in {"descricao", "descricao"}:
                    return col
        return None

    def _find_month_blocks(self, worksheet) -> list[dict[str, Any]]:
        month_blocks: list[dict[str, Any]] = []

        merged_ranges = list(worksheet.merged_cells.ranges)
        for merged in merged_ranges:
            start_cell = worksheet.cell(row=merged.min_row, column=merged.min_col)
            label = self._normalize_label(start_cell.value)
            if not self._is_month_column_name(label):
                continue

            total_col = None
            for row in range(merged.min_row, min(merged.min_row + 4, worksheet.max_row) + 1):
                for col in range(merged.max_col + 1, min(merged.max_col + 80, worksheet.max_column) + 1):
                    text = self._normalize_label(worksheet.cell(row=row, column=col).value)
                    if "total" in text:
                        total_col = col
                        break
                if total_col is not None:
                    break

            month_blocks.append(
                {
                    "month": start_cell.value,
                    "month_sort": self._month_sort_key(str(start_cell.value)),
                    "header_row": merged.min_row,
                    "start_col": merged.min_col,
                    "end_col": merged.max_col,
                    "total_col": total_col,
                }
            )

        month_blocks.sort(key=lambda block: (block["month_sort"], block["header_row"], block["start_col"]))
        dedup: OrderedDict[str, dict[str, Any]] = OrderedDict()
        for block in month_blocks:
            month_name = str(block["month"]).strip().upper()
            if month_name not in dedup:
                dedup[month_name] = block
        return list(dedup.values())

    def _pick_target_description_patterns(self, metric_tag: str) -> tuple[list[str], list[str]]:
        if metric_tag == "__finance_receitas":
            return ["(+) receita sobre vendas", "receitas"], ["receita", "venda"]
        if metric_tag == "__finance_despesas":
            return ["(+/-) outras receitas ou despesas", "despesas"], ["despesa", "outras receitas ou despesas", "gasto", "custo"]
        if metric_tag == "__finance_resultado":
            return ["(=) resultado liquido", "superavit/deficit", "superavit / deficit", "(=) fluxo de caixa livre - (fcl)"], ["resultado", "superavit", "deficit", "fcl", "fluxo de caixa"]
        if metric_tag == "__finance_saldos":
            return ["saldos", "(=) saldo mes anterior"], ["saldo", "caixa"]
        return [], []

    def _find_row_for_metric(self, worksheet, description_col: int, metric_tag: str, scan_start: int, scan_end: int) -> int | None:
        preferred, fallback = self._pick_target_description_patterns(metric_tag)
        normalized_preferred = [self._normalize_name(self._normalize_label(label)) for label in preferred]
        normalized_fallback = [self._normalize_name(self._normalize_label(label)) for label in fallback]

        candidates: list[tuple[int, int]] = []
        for row in range(scan_start, scan_end + 1):
            text = self._normalize_label(worksheet.cell(row=row, column=description_col).value)
            if not text:
                continue

            normalized = self._normalize_name(text)
            if normalized in normalized_preferred:
                return row

            score = sum(1 for token in normalized_fallback if token and token in normalized)
            if score:
                candidates.append((score, row))

        if not candidates:
            return None

        candidates.sort(key=lambda item: (-item[0], item[1]))
        return candidates[0][1]

    def _sum_month_cells(self, worksheet, row: int, start_col: int, end_col: int) -> float:
        total = 0.0
        for col in range(start_col, end_col + 1):
            parsed = self._parse_numeric_value(worksheet.cell(row=row, column=col).value)
            if parsed is not None:
                total += parsed
        return total

    def _extract_wide_finance_from_workbook(
        self,
        file_uuid: str,
        sheet_name: str,
        metric_tag: str,
    ) -> list[dict[str, Any]]:
        workbook_path = self._resolve_original_workbook_path(file_uuid)
        if not workbook_path or workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            return []

        workbook = load_workbook(workbook_path, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                return []

            worksheet = workbook[sheet_name]
            description_col = self._find_description_col_index(worksheet)
            if description_col is None:
                return []

            month_blocks = self._find_month_blocks(worksheet)
            if not month_blocks:
                return []

            results: list[dict[str, Any]] = []
            for block in month_blocks:
                scan_start = block["header_row"] + 1
                next_header_row = worksheet.max_row
                for other in month_blocks:
                    if other["header_row"] > block["header_row"]:
                        next_header_row = min(next_header_row, other["header_row"] - 1)
                scan_end = next_header_row

                row = self._find_row_for_metric(worksheet, description_col, metric_tag, scan_start, scan_end)
                if row is None:
                    continue

                total_value = None
                if block["total_col"] is not None:
                    total_value = self._parse_numeric_value(worksheet.cell(row=row, column=block["total_col"]).value)

                if total_value is None:
                    total_value = self._sum_month_cells(worksheet, row, block["start_col"], block["end_col"])

                month_label = str(block["month"]).strip().upper()
                results.append({"__month": month_label, metric_tag: float(total_value or 0.0)})

            return results
        finally:
            workbook.close()


analytics_service = AnalyticsService()
