"""Tests for deterministic ingestion behavior."""
from __future__ import annotations

import tempfile
from pathlib import Path

import polars as pl
import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.core.exceptions import UnsupportedFormatException
from app.services.ingestion import ingestion_service


class TestIngestionService:
    """Core ingestion tests for Prompt 1.2."""

    @pytest.mark.asyncio
    async def test_calculate_sha256(self):
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as file_obj:
            file_obj.write("a,b\n1,2\n")
            path = file_obj.name

        result = await ingestion_service._calculate_sha256(path)
        assert len(result) == 64
        Path(path).unlink(missing_ok=True)

    def test_detect_string_leading_zeros_stays_string(self):
        series = pl.Series("codigo", ["00123", "00456", "00001"])
        assert ingestion_service._detect_column_type(series) == "string"

    def test_detect_percentage_type(self):
        series = pl.Series("taxa", ["10%", "20%", "30%"])
        assert ingestion_service._detect_column_type(series) == "percentage"

    def test_detect_currency_type(self):
        series = pl.Series("valor", ["R$ 10,00", "R$ 20,00", "R$ 30,00"])
        assert ingestion_service._detect_column_type(series) == "currency"

    def test_effective_sheet_bounds_ignore_formatted_padding_cells(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "SheetA"
        worksheet.append(["valor", "descricao"])
        worksheet.append([1, "A"])
        worksheet.append([2, "B"])
        worksheet["Z50"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")

        rows, columns = ingestion_service._get_effective_sheet_bounds(worksheet)

        assert rows == 2
        assert columns == 2

    @pytest.mark.asyncio
    async def test_unsupported_format(self):
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as file_obj:
            path = file_obj.name

        with pytest.raises(UnsupportedFormatException):
            await ingestion_service.process_upload(path, "invalid.pdf")

        Path(path).unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_strict_mode_marks_inconsistent_without_blocking(self, monkeypatch):
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / "multi.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "SheetA"
            ws.append(["valor"])
            ws.append([1])
            wb.save(file_path)

            async def fake_read_primary(*_args, **_kwargs):
                return {"SheetA": pl.DataFrame({"valor": [1]})}

            async def fake_read_secondary(*_args, **_kwargs):
                return {"SheetA": {"max_row": 2, "max_column": 1, "merged_cells_count": 0, "formula_count": 0, "sheet_hash": "x"}}

            def fake_compare(*_args, **_kwargs):
                return [{"type": "row_count_mismatch", "severity": "high", "sheet": "SheetA"}]

            async def fake_persist(*_args, **_kwargs):
                return Path(tmp_dir) / "analytics.db"

            monkeypatch.setattr(ingestion_service, "_read_primary", fake_read_primary)
            monkeypatch.setattr(ingestion_service, "_read_secondary_xlsx", fake_read_secondary)
            monkeypatch.setattr(ingestion_service, "_compare_readings", fake_compare)
            monkeypatch.setattr(ingestion_service, "_persist_to_duckdb", fake_persist)

            result = await ingestion_service.process_upload(str(file_path), "multi.xlsx", strict_mode=True)

            assert result["status"] == "inconsistent"
            assert result["strict_mode"] is True
            assert result["integrity_report"]["strict_mode_blocked"] is True
            assert result["duckdb_path"]

    @pytest.mark.asyncio
    async def test_non_strict_mode_marks_inconsistent(self, monkeypatch):
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / "multi.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "SheetA"
            ws.append(["valor"])
            ws.append([1])
            wb.save(file_path)

            async def fake_read_primary(*_args, **_kwargs):
                return {"SheetA": pl.DataFrame({"valor": [1]})}

            async def fake_read_secondary(*_args, **_kwargs):
                return {"SheetA": {"max_row": 1, "max_column": 1, "merged_cells_count": 0, "formula_count": 0, "sheet_hash": "x"}}

            def fake_compare(*_args, **_kwargs):
                return [{"type": "row_count_mismatch", "severity": "medium", "sheet": "SheetA"}]

            async def fake_persist(*_args, **_kwargs):
                return Path(tmp_dir) / "analytics.db"

            monkeypatch.setattr(ingestion_service, "_read_primary", fake_read_primary)
            monkeypatch.setattr(ingestion_service, "_read_secondary_xlsx", fake_read_secondary)
            monkeypatch.setattr(ingestion_service, "_compare_readings", fake_compare)
            monkeypatch.setattr(ingestion_service, "_persist_to_duckdb", fake_persist)

            result = await ingestion_service.process_upload(str(file_path), "multi.xlsx", strict_mode=False)

            assert result["status"] == "inconsistent"
            report = result["integrity_report"]
            assert "engine_divergences" in report
            assert "columns_by_type" in report
            assert "sheet_hashes" in report
